"""프로젝트별 분석·설계 문서(엑셀)를 JSON으로 뽑는다.
엑셀 캡처를 붙이는 대신 사이트 서식으로 다시 그리기 위한 원본 데이터.

시트마다 제목 줄 수·헤더 단 수가 달라서 자동 감지를 믿지 않고 시트별로 지정한다.
`hdr`는 헤더로 쓸 행 번호(빈 행 제거 후 기준). 두 개면 2단 헤더를 셀 단위로 합친다.
문서 중간에 반복되는 그룹 머리행(예: `[회원/인증] users`)은 표의 구분 띠로 따로 넘긴다."""
import colorsys, json, os, re, sys
import openpyxl

D = r"C:\Users\USER\Downloads"

# (프로젝트, 파일, [(키, 시트명, 라벨, 헤더행)])
SPEC = [
    ("cogi", "404(COGI)_분석, 설계 통합본 최종.xlsx", [
        ("req",   "요구사항정의서", "요구사항 정의서", [0]),
        ("func",  "기능 정의서",    "기능 정의서",     [0, 1]),
        ("wbs",   "WBS 최종",      "WBS",            [1, 2], True),
        ("api",   "API 명세서",    "API 명세서",      [0]),
        ("table", "테이블 정의서",  "테이블 정의서",    [2]),
    ]),
    ("tl", "TripLinker 분석,설계,개발.xlsx", [
        ("req",   "요구사항 리스트 최종", "요구사항 정의서",     [1]),
        ("func",  "기능리스트",          "기능 정의서",        [1, 2]),
        ("wbs",   "WBS 최종",           "WBS",               [0, 1], False, "구분"),  # 일정은 ● 문자, 색은 배경 워시
        ("api",   "API정의서",           "API 명세서",         [1]),
        ("table", "테이블 정의서",        "테이블 정의서",       [2]),
        ("test",  "테스트 케이스(정상연)", "테스트 케이스 (담당분)", [0, 1]),
    ]),
    ("om", "오몽 프로젝트 설게 분석 개발 문서.xlsx", [
        ("req",   "1.요구사항정의서", "요구사항 정의서", [2]),
        ("func",  "2.기능정의서",     "기능 정의서",    [2]),
        ("wbs",   "5.WBS_3박4일",    "WBS",           [2], True),
        ("api",   "4.API정의서",      "API 명세서",     [3]),
        ("table", "3.테이블정의서",    "테이블 정의서",   [3]),
        ("ai",    "6.AI활용_로그",    "AI 활용 로그",   [3]),
    ]),
]


def clean(v):
    if v is None:
        return ""
    s = str(v).replace("\r\n", "\n").strip()
    return " ".join(s.split()) if "\n" not in s else s


def wide_merge_rows(ws, span=3):
    """가로로 3칸 이상 병합된 행 번호(1-base). 엑셀에서 그룹 머리행은 이렇게 병합돼 있다 —
    값의 개수로 추측하는 것보다 문서 구조를 그대로 읽는 쪽이 정확하다."""
    out = set()
    for m in ws.merged_cells.ranges:
        if m.max_col - m.min_col + 1 >= span:
            for r in range(m.min_row, m.max_row + 1):
                out.add(r)
    return out


def grid(ws):
    """(행 값, 원본 행번호) — 병합 정보와 맞추려면 원본 행번호가 필요하다."""
    rows, nums = [], []
    for n, r in enumerate(ws.iter_rows(values_only=True), start=1):
        vals = [clean(c) for c in r]
        if any(vals):
            rows.append(vals)
            nums.append(n)
    if not rows:
        return [], []
    width = max(len(r) for r in rows)
    return [r + [""] * (width - len(r)) for r in rows], nums


def merge_header(rows, idx):
    top = rows[idx[0]]
    if len(idx) == 1:
        return list(top)
    sub = rows[idx[1]]
    out = []
    for i in range(len(top)):
        a, b = top[i], sub[i] if i < len(sub) else ""
        out.append((a + " " + b).strip() if a and b else (a or b))
    return out


def bar_map(ws):
    """값은 없는데 배경색만 있는 셀 = 간트 바. {(행,열): '#rrggbb'}
    WBS는 일정을 셀 색으로 표시해서, 글자만 읽으면 '언제 했는지'가 통째로 사라진다."""
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                continue
            f = c.fill
            if not (f and f.patternType):
                continue
            rgb = getattr(f.start_color, "rgb", None)
            if not rgb or not isinstance(rgb, str) or len(rgb) != 8 or rgb == "00000000":
                continue
            if rgb[2:].upper() in ("FFFFFF", "F3F3F3"):      # 흰색·연회색은 장식
                continue
            out[(c.row, c.column)] = "#" + rgb[2:].lower()
    return out


def extract(ws, hdr, use_bars=False, cut_at=None):
    rows, nums = grid(ws)
    if not rows:
        return None
    merged = wide_merge_rows(ws)
    bars = bar_map(ws) if use_bars else {}
    head = merge_header(rows, hdr)
    start = max(hdr) + 1
    body, body_nums = rows[start:], nums[start:]

    # 헤더가 비어 있고 값도 거의 없는 열은 버린다 (엑셀의 장식용 빈 열)
    keep = [i for i in range(len(head))
            if head[i] or sum(1 for r in body if i < len(r) and r[i]) >= 3]
    head = [head[i] for i in keep]
    # 표 오른쪽에 붙어 있는 별도 범례 블록은 잘라낸다 (본 표와 무관한 열)
    if cut_at:
        for i, h in enumerate(head):
            if h.strip() == cut_at:
                keep, head = keep[:i], head[:i]
                break
    norm = [h.strip() for h in head]

    out, legend = [], []
    for r, n in zip(body, body_nums):
        cells = [r[i] if i < len(r) else "" for i in keep]
        filled = [c for c in cells if c]
        rowbars = {str(j): bars[(n, keep[j] + 1)] for j in range(len(keep))
                   if (n, keep[j] + 1) in bars}
        if not filled and not rowbars:
            continue
        if [c.strip() for c in cells] == norm:
            continue                                   # 표 중간에 반복되는 헤더 행은 버린다
        if all(c in ("(작성)", "-") for c in filled):
            continue                                   # 채우지 않은 템플릿 행은 버린다
        # 표 아래 붙은 색 범례 행 (`담당 | 정상연 (팀장) | 이수환 | …`) — 표가 아니라 범례로 넘긴다
        if use_bars and "담당" in filled and len(filled) >= 3:
            seen = False
            for j, col in enumerate(keep):
                v = cells[j]
                if v == "담당":
                    seen = True
                    continue
                if not (seen and v):
                    continue
                cell = ws.cell(row=n, column=col + 1)
                rgb = getattr(cell.fill.start_color, "rgb", None) if cell.fill and cell.fill.patternType else None
                if rgb and isinstance(rgb, str) and len(rgb) == 8 and rgb[2:].upper() not in ("FFFFFF", "F3F3F3"):
                    legend.append({"who": v, "color": "#" + rgb[2:].lower()})
            if legend:
                continue
        if n in merged and len(filled) <= max(4, len(head) // 2) and not rowbars:
            out.append({"g": "  ".join(filled)})       # 그룹 머리행 → 구분 띠
            continue
        out.append({"c": cells, "b": rowbars} if rowbars else cells)

    got = {"head": head, "rows": out}
    if legend:
        got["legend"] = legend
    return got


TODAY = "2026-07-30"     # 기준일. 스크립트가 날짜를 추측하지 않게 명시한다.


def freshen(proj, key, doc):
    """WBS를 기준일 시점으로 최신화한다.
    - 일정 표시(●)가 빠진 행은 같은 업무 묶음의 표시 주차를 따라 채운다
    - 계획 일자가 이미 지난 행은 상태를 완료로 올린다
    - COGI는 상태 열이 없어 주차 바 위치로 상태를 도출해 붙인다 (진행 중이라 8월 W5는 예정 유지)
    문서를 손댄 것이므로 뷰어 하단에 최신화 사실을 표시한다."""
    if key != "wbs":
        return
    head, rows = doc["head"], doc["rows"]
    data = [r for r in rows if not (isinstance(r, dict) and "g" in r)]

    def cells(r):
        return r["c"] if isinstance(r, dict) else r

    weeks = [i for i, h in enumerate(head) if re.search(r"[wW]\d", h)]
    status = next((i for i, h in enumerate(head) if "상태" in h or "현재" in h), None)

    if proj == "tl":
        # 표시가 없는 행을 앞뒤에서 그대로 복사하면 한 주차에 다 몰려 일정으로 안 읽힌다.
        # 문서 자체의 구조(단계 → 업무 묶음)를 기준으로 계단식으로 배치한다.
        #   · 원본에 이미 있는 표시는 기준점으로 삼아 그대로 둔다
        #   · 표시 없는 묶음은 그 단계의 주차 구간에 순서대로 나눠 넣는다
        #   · 뒤 묶음이 앞 묶음보다 앞선 주차로 가지 않게 단조 증가를 강제한다
        PHASE = {"설계": (0, 2), "구현": (3, 6), "테스트": (6, 6), "배포": (7, 7), "마무리": (7, 7)}
        groups, order, phase, grp = {}, [], "", ""
        for r in data:
            c = cells(r)
            phase = c[0] or phase                    # 병합으로 빈 칸은 위 값을 이어받는다
            grp = c[1] or grp
            key = (phase, grp)
            if key not in groups:
                groups[key] = []
                order.append(key)
            here = next((i for i in weeks if "●" in c[i]), None)
            groups[key].append((r, here))

        assigned, last = {}, -1
        by_phase = {}
        for key in order:
            by_phase.setdefault(key[0], []).append(key)
        for ph, keys in by_phase.items():
            lo, hi = PHASE.get(ph, (0, len(weeks) - 1))
            for n, key in enumerate(keys):
                # 앵커는 열 번호이므로 주차 위치(0-base)로 바꿔서 쓴다
                anchors = [weeks.index(h) for _, h in groups[key] if h is not None]
                if anchors:
                    w = min(anchors)                 # 원본 표시가 있으면 그것이 기준
                else:
                    span = hi - lo
                    w = lo + (round(span * n / max(1, len(keys) - 1)) if span else 0)
                w = max(w, last)                     # 단조 증가
                assigned[key] = w
                last = w

        for key in order:
            w = assigned[key]
            for r, here in groups[key]:
                if here is None:
                    cells(r)[weeks[w] if w < len(weeks) else weeks[-1]] = "●"
        # 5~6월 일정이라 기준일 기준 전부 지난 시점
        if status is not None:
            for r in data:
                c = cells(r)
                if c[status] in ("", "예정", "진행"):
                    c[status] = "완료"
        doc["note"] = "일정 표시(●)와 완료 상태는 " + TODAY + " 기준으로 최신화했습니다. 계획 일자(5~6월)가 모두 지나 완료로 반영했습니다."

    elif proj == "om":
        if status is not None:
            for r in data:
                c = cells(r)
                if c[status] in ("", "예정", "진행"):
                    c[status] = "완료"
        doc["note"] = "상태는 " + TODAY + " 기준으로 최신화했습니다. 계획 일자(6/30~7/3)가 모두 지나 완료로 반영했습니다."

    elif proj == "cogi":
        # 주차 헤더의 종료일로 판별한다. 종료일이 지났으면 완료, 기준일이 걸친 주차는 진행, 그 뒤는 예정.
        # `7/1~10`(종료월 생략)과 `7/27~8/2`(월 넘어감) 두 형식을 모두 읽는다
        ends = {}
        for i in weeks:
            m = re.search(r"(\d+)/(\d+)\s*~\s*(?:(\d+)/)?(\d+)", head[i])
            if m:
                mon = int(m.group(3) or m.group(1))
                ends[i] = "2026-%02d-%02d" % (mon, int(m.group(4)))
        done = {i for i, e in ends.items() if e < TODAY}
        rest = sorted(i for i in ends if i not in done)
        current = rest[0] if rest else None
        head.append("상태")
        for r in rows:
            if isinstance(r, dict) and "g" in r:
                continue
            c = cells(r)
            bars = sorted(int(k) for k in (r.get("b", {}) if isinstance(r, dict) else {}))
            if not bars:
                c.append("")
                continue
            last = bars[-1]
            c.append("완료" if last in done else ("진행" if last == current else "예정"))
        # 범례 행은 바와 같은 색상을 연하게 칠해 뒀다. 색상(hue)이 가장 가까운 바 색으로 맞춰
        # 범례 칩과 표의 바가 같은 색으로 보이게 한다.
        legend = doc.get("legend") or []
        barcols = []
        for r in rows:
            for v in (r.get("b", {}) if isinstance(r, dict) else {}).values():
                if v not in barcols:
                    barcols.append(v)
        if legend and barcols:
            def hue(h):
                r_, g_, b_ = (int(h[i:i + 2], 16) / 255 for i in (1, 3, 5))
                return colorsys.rgb_to_hsv(r_, g_, b_)[0]
            for item in legend:
                item["color"] = min(barcols, key=lambda b: abs(hue(b) - hue(item["color"])))

        doc["note"] = ("상태는 " + TODAY + " 기준으로 주차 일정에서 도출했습니다. "
                       "7월 W1~W3은 완료, 기준일이 속한 W4는 진행, 8월 W5는 예정입니다.")


def main(out_dir):
    report = []
    for proj, fname, sheets in SPEC:
        wb = openpyxl.load_workbook(os.path.join(D, fname), data_only=True)
        docs = {}
        for row in sheets:
            key, sheet, label, hdr = row[:4]
            use_bars = len(row) > 4 and row[4]
            cut_at = row[5] if len(row) > 5 else None
            if sheet not in wb.sheetnames:
                report.append("MISS %s/%s (%s)" % (proj, key, sheet))
                continue
            got = extract(wb[sheet], hdr, use_bars, cut_at)
            if not got:
                report.append("EMPTY %s/%s" % (proj, key))
                continue
            docs[key] = {"label": label, "sheet": sheet, "file": fname, **got}
            freshen(proj, key, docs[key])      # WBS를 기준일 시점으로 최신화
            bands = sum(1 for r in got["rows"] if isinstance(r, dict) and "g" in r)
            gantt = sum(1 for r in got["rows"] if isinstance(r, dict) and r.get("b"))
            report.append("OK   %s/%-5s %-18s %3d행 (구분띠 %2d · 간트바 %2d) x %2d열"
                          % (proj, key, label, len(got["rows"]), bands, gantt, len(got["head"])))
        wb.close()
        p = os.path.join(out_dir, "docs-%s.js" % proj)
        open(p, "w", encoding="utf-8").write(
            "window.PROJECT_DOCS = " + json.dumps(docs, ensure_ascii=False, separators=(",", ":")) + ";\n")
        report.append("  -> docs-%s.js  %.0f KB" % (proj, os.path.getsize(p) / 1024))
    print("\n".join(report))


if __name__ == "__main__":
    main(sys.argv[1])
